# https://linuxhint.com/read-excel-file-python/
# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import openpyxl
import logging

from text_grid import TextGrid

# -----------------------------------------------------------------------------

def read_xlsx_sheet(path: str) -> TextGrid:
    """
    Reads entire sheet 0
    """

    logging.info(f"Reading file '{path}'")
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook(path)
    sheet = wookbook.active
    tg = TextGrid()

    # Iterate the loop to read the cell values
    for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
        row_cells = []
        for cell in row:
            if cell is None:
                cell = ""
            elif type(cell) is float or type(cell) is int:
                cell = repr(cell)
            row_cells.append(cell.strip())

        # ignore rows with empty cell 'A'
        if row_cells and row_cells[0] != "":
            tg.rows.append(row_cells)

    tg.nrows = len(tg.rows)
    tg.ncols = sheet.max_column
    tg.align_number_of_columns()
    return tg
